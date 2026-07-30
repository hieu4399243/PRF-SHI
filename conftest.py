"""Sinh báo cáo kết quả unit test ra Excel/CSV để dán vào Google Sheets.

    ./.venv/bin/python -m pytest --excel-report
    ./.venv/bin/python -m pytest --excel-report=duong/dan/khac.xlsx

Mỗi lần chạy ghi ra 2 file cạnh nhau:
  - .xlsx  -> mở bằng Excel / Numbers, có màu và bộ lọc sẵn.
  - .csv   -> Google Sheets: File > Import > Upload > Replace current sheet.

Không truyền --excel-report thì pytest chạy y như cũ, không sinh file nào.
"""

import csv
import datetime
import sys
from pathlib import Path

import pytest

DEFAULT_REPORT = "reports/ket_qua_unit_test.xlsx"

# Tên tiếng Việt cho từng file test — thành cột "Nhóm chức năng" trong báo cáo.
NHOM_CHUC_NANG = {
    "test_app_admin.py": "API quản trị (admin)",
    "test_app_hardening.py": "Bảo mật (rate limit, kích thước body, session)",
    "test_app_ics.py": "API xuất lịch .ics",
    "test_booking.py": "Đặt lịch khám",
    "test_calendar_ics.py": "Sinh file lịch .ics",
    "test_chatbot_audit.py": "Chatbot — nhật ký audit",
    "test_chatbot_flex.py": "Chatbot — hội thoại linh hoạt",
    "test_chatbot_guardrail.py": "Chatbot — guardrail",
    "test_chatbot_session_lock.py": "Chatbot — khoá phiên",
    "test_chatbot_sessions.py": "Chatbot — quản lý phiên",
    "test_push.py": "Thông báo đẩy (push)",
    "test_reminder_worker.py": "Worker nhắc lịch hẹn",
    "test_safety.py": "An toàn y tế (cấp cứu, tự chẩn đoán)",
    "test_storage.py": "Lưu trữ (JSON / Postgres)",
    "test_triage_llm.py": "Phân loại triệu chứng bằng LLM",
    "test_triage_negation.py": "Phân loại triệu chứng — câu phủ định",
}

COLUMNS = [
    ("STT", 6),
    ("Mã test", 10),
    ("Nhóm chức năng", 34),
    ("File", 28),
    ("Tên test", 44),
    ("Bộ dữ liệu (tham số)", 30),
    ("Mô tả / Kết quả mong đợi", 52),
    ("Trạng thái", 12),
    ("Thời gian (s)", 13),
    ("Chi tiết lỗi", 60),
]

# Màu nền cho cột "Trạng thái".
MAU_TRANG_THAI = {
    "PASS": "C6EFCE",
    "FAIL": "FFC7CE",
    "ERROR": "FFC7CE",
    "SKIP": "FFEB9C",
    "XFAIL": "FFEB9C",
    "XPASS": "FFEB9C",
}

# pytest_runtest_logreport chỉ nhận `report`, không có đường nào lần ra config,
# nên giữ tham chiếu ở đây (mỗi tiến trình pytest chỉ có 1 config).
_CONFIG = None


def pytest_configure(config):
    global _CONFIG
    _CONFIG = config


def pytest_addoption(parser):
    parser.addoption(
        "--excel-report",
        nargs="?",
        const=DEFAULT_REPORT,
        default=None,
        metavar="PATH",
        help=f"Xuất kết quả test ra Excel + CSV (mặc định {DEFAULT_REPORT}).",
    )


def _mo_ta_case(item):
    """Mô tả bộ dữ liệu của test parametrize, ví dụ: text=tôi không bị đau răng.

    Không dùng `callspec.id` vì pytest escape chữ có dấu thành \\xf4, \\u1ecb...
    -> báo cáo tiếng Việt đọc không ra.
    """
    spec = getattr(item, "callspec", None)
    if spec is None:
        return ""
    phan = []
    for ten, gia_tri in spec.params.items():
        s = str(gia_tri).replace("\n", " ")
        phan.append(f"{ten}={s[:77] + '...' if len(s) > 80 else s}")
    return "; ".join(phan) or spec.id


def _dong_dau_docstring(obj):
    """Lấy dòng đầu tiên có nội dung của docstring, dùng làm mô tả test."""
    doc = getattr(obj, "__doc__", None) or ""
    for line in doc.strip().splitlines():
        if line.strip():
            return line.strip()
    return ""


@pytest.hookimpl(trylast=True)
def pytest_collection_modifyitems(session, config, items):
    """Ghi nhận thứ tự + mô tả của từng test ngay khi collect xong.

    Làm ở đây vì hook báo kết quả (pytest_runtest_logreport) chỉ nhận được
    `report`, không có `item` để đọc docstring.
    """
    if not config.getoption("--excel-report"):
        return
    ket_qua = {}
    for stt, item in enumerate(items, start=1):
        mo_ta = _dong_dau_docstring(getattr(item, "function", None))
        if not mo_ta:  # test không có docstring -> mượn mô tả chung của file
            mo_ta = _dong_dau_docstring(getattr(item, "module", None))
        ten_file = Path(str(item.path)).name
        ket_qua[item.nodeid] = {
            "stt": stt,
            "ma": f"TC-{stt:03d}",
            "nhom": NHOM_CHUC_NANG.get(ten_file, ten_file),
            "file": ten_file,
            "ten": getattr(item, "originalname", None) or item.name,
            "case": _mo_ta_case(item),
            "mo_ta": mo_ta,
            "trang_thai": "SKIP",
            "thoi_gian": 0.0,
            "loi": "",
        }
    config._excel_ket_qua = ket_qua


def pytest_runtest_logreport(report):
    """Gom kết quả của cả 3 giai đoạn setup/call/teardown cho mỗi test."""
    ket_qua = getattr(_CONFIG, "_excel_ket_qua", None)
    if not ket_qua:
        return
    dong = ket_qua.get(report.nodeid)
    if dong is None:
        return

    dong["thoi_gian"] += report.duration

    if report.when == "call":
        if hasattr(report, "wasxfail"):
            dong["trang_thai"] = "XPASS" if report.passed else "XFAIL"
        elif report.passed:
            dong["trang_thai"] = "PASS"
        elif report.failed:
            dong["trang_thai"] = "FAIL"
        else:
            dong["trang_thai"] = "SKIP"
    elif report.failed:
        # Hỏng ở setup/teardown là lỗi hạ tầng test, không phải test fail.
        dong["trang_thai"] = "ERROR"
    elif report.when == "setup" and report.skipped:
        dong["trang_thai"] = "SKIP"

    if report.failed and not dong["loi"]:
        chi_tiet = report.longreprtext.strip()
        dong["loi"] = chi_tiet[-1500:] if len(chi_tiet) > 1500 else chi_tiet


def pytest_sessionfinish(session, exitstatus):
    duong_dan = session.config.getoption("--excel-report")
    if not duong_dan:
        return
    ket_qua = getattr(session.config, "_excel_ket_qua", None)
    if not ket_qua:
        return

    dich = Path(duong_dan)
    if dich.suffix.lower() != ".xlsx":
        dich = dich.with_suffix(".xlsx")
    dich.parent.mkdir(parents=True, exist_ok=True)

    dong_list = sorted(ket_qua.values(), key=lambda d: d["stt"])
    tom_tat = _tinh_tom_tat(dong_list)

    duong_dan_csv = dich.with_suffix(".csv")
    _ghi_csv(duong_dan_csv, dong_list)
    da_ghi = [duong_dan_csv]
    try:
        _ghi_xlsx(dich, dong_list, tom_tat)
        da_ghi.insert(0, dich)
    except ImportError:
        print("\n[excel-report] Chưa cài openpyxl -> chỉ ghi được CSV. "
              "Cài bằng: ./.venv/bin/python -m pip install openpyxl", file=sys.stderr)

    print("\n[excel-report] " + " | ".join(
        f"{k}: {v}" for k, v in tom_tat.items() if k != "Ngày chạy"))
    for f in da_ghi:
        print(f"[excel-report] Đã ghi: {f}")


def _tinh_tom_tat(dong_list):
    dem = {}
    for d in dong_list:
        dem[d["trang_thai"]] = dem.get(d["trang_thai"], 0) + 1
    tong = len(dong_list)
    so_pass = dem.get("PASS", 0)
    return {
        "Ngày chạy": datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "Tổng số test": tong,
        "Đạt (PASS)": so_pass,
        "Không đạt (FAIL)": dem.get("FAIL", 0),
        "Lỗi (ERROR)": dem.get("ERROR", 0),
        "Bỏ qua (SKIP)": dem.get("SKIP", 0) + dem.get("XFAIL", 0) + dem.get("XPASS", 0),
        "Tỉ lệ đạt": f"{so_pass / tong * 100:.1f}%" if tong else "0%",
        "Tổng thời gian": f"{sum(d['thoi_gian'] for d in dong_list):.2f}s",
        "Phiên bản Python": sys.version.split()[0],
        "Phiên bản pytest": pytest.__version__,
    }


def _hang_du_lieu(d):
    return [d["stt"], d["ma"], d["nhom"], d["file"], d["ten"], d["case"],
            d["mo_ta"], d["trang_thai"], round(d["thoi_gian"], 4), d["loi"]]


def _ghi_csv(duong_dan, dong_list):
    # utf-8-sig để Excel/Google Sheets đọc đúng dấu tiếng Việt.
    with open(duong_dan, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([ten for ten, _ in COLUMNS])
        for d in dong_list:
            w.writerow(_hang_du_lieu(d))


def _ghi_xlsx(duong_dan, dong_list, tom_tat):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    ws = wb.active
    ws.title = "Tổng hợp"
    ws["A1"] = "BÁO CÁO KẾT QUẢ UNIT TEST"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Dự án: AI Health Assistant (SHI)"
    for i, (khoa, gia_tri) in enumerate(tom_tat.items(), start=4):
        ws[f"A{i}"] = khoa
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = gia_tri
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 26

    # Thống kê theo từng nhóm chức năng.
    dong_bat_dau = len(tom_tat) + 6
    ws[f"A{dong_bat_dau}"] = "Theo nhóm chức năng"
    ws[f"A{dong_bat_dau}"].font = Font(bold=True, size=12)
    for tieu_de, cot in (("Nhóm chức năng", "A"), ("Tổng", "B"), ("Đạt", "C"), ("Không đạt", "D")):
        o = ws[f"{cot}{dong_bat_dau + 1}"]
        o.value = tieu_de
        o.font = Font(bold=True)
    theo_nhom = {}
    for d in dong_list:
        muc = theo_nhom.setdefault(d["nhom"], {"tong": 0, "dat": 0, "hong": 0})
        muc["tong"] += 1
        if d["trang_thai"] == "PASS":
            muc["dat"] += 1
        elif d["trang_thai"] in ("FAIL", "ERROR"):
            muc["hong"] += 1
    for i, (nhom, muc) in enumerate(sorted(theo_nhom.items()), start=dong_bat_dau + 2):
        ws[f"A{i}"] = nhom
        ws[f"B{i}"] = muc["tong"]
        ws[f"C{i}"] = muc["dat"]
        ws[f"D{i}"] = muc["hong"]
    ws.column_dimensions["A"].width = 40

    ws2 = wb.create_sheet("Chi tiết")
    tieu_de_fill = PatternFill("solid", fgColor="1F4E78")
    for col, (ten, do_rong) in enumerate(COLUMNS, start=1):
        o = ws2.cell(row=1, column=col, value=ten)
        o.font = Font(bold=True, color="FFFFFF")
        o.fill = tieu_de_fill
        o.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws2.column_dimensions[get_column_letter(col)].width = do_rong

    for r, d in enumerate(dong_list, start=2):
        for c, gia_tri in enumerate(_hang_du_lieu(d), start=1):
            o = ws2.cell(row=r, column=c, value=gia_tri)
            o.alignment = Alignment(vertical="top", wrap_text=c in (7, 10))
        o_trang_thai = ws2.cell(row=r, column=8)
        o_trang_thai.alignment = Alignment(horizontal="center", vertical="top")
        mau = MAU_TRANG_THAI.get(d["trang_thai"])
        if mau:
            o_trang_thai.fill = PatternFill("solid", fgColor=mau)
            o_trang_thai.font = Font(bold=True)

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(dong_list) + 1}"

    wb.save(duong_dan)
